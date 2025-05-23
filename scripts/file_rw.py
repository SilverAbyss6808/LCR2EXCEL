
# this file is gonna be just for file interactions

import openpyxl as opxl
import pypdf
import string
import os
import data_processing as dp
import visual_formatting as vf


def read_input_file(path: string, filetype: string):
    # check if the path exists and is a valid pdf/xlsx
    try:
        if os.path.splitext(path)[1] != filetype:
            raise TypeError(f'not a {filetype} file')

        if os.path.exists(path):
            if filetype == '.pdf':
                read_pdf(path)
            elif filetype == '.xlsx':
                pass
        else:
            raise FileNotFoundError(f'File {path} does not exist. Please try again.')

        return True  # returns True if the file exists

    except ValueError as ve:  # file's not open probably
        print(f'Error: {ve}')
    except NameError as ne:  # undefined variables, etc
        print(f'Error: {ne}')
    except FileNotFoundError as nf:  # no file at specified path
        print(f'Error: {nf}')
    except TypeError as te:  # provided file is not pdf or xlsx, depending on type
        print(f'Error: {te}')

    return False  # only executes if an error occurs


def validate_proposed_filepath(path: string, filetype: string):
    directory = os.path.dirname(path)

    if directory is not None:
        if not os.path.isfile(path):
            if filetype == '.xlsx':
                return True
        else:
            # todo: ask if it's ok to overwrite the file rather than just saying fuck you
            print('Directory is valid, but file already exists. Please enter a path with a new filename.')
            return False
    else:
        print(f'Path {path} is not in a valid directory.')
        return False


def read_pdf(path: string):
    # path has already been verified so its ok
    data: string = ''

    reader = pypdf.PdfReader(path)
    pages = reader.pages
    for page in pages:
        data += page.extract_text()

    jobs_list: list[dp.Job] = dp.format_pdf_data_as_job(data)
    return jobs_list


def read_excel(input_excel_path: string):
    excel_jobs: list[string] = []

    workbook = opxl.load_workbook(input_excel_path)
    active_sheet = workbook.active

    rows: list = list(active_sheet.iter_rows(min_row=2, values_only=True))
    num_rows = active_sheet.max_row - 1

    for i in range(0, num_rows):  # i is the index jsyk
        row_num = i + 2

        current_row = (rows[i])
        dim_data = ()
        add_to_dd = ''

        if active_sheet.row_dimensions[row_num].outlineLevel > 0:  # these are rows that are grouped together
            add_to_dd += 'g'

        if active_sheet.row_dimensions[row_num].hidden:  # rows that are hidden
            add_to_dd += 'h'

        dim_data += (add_to_dd,)
        current_row = current_row + dim_data
        excel_jobs.append(current_row)

    excel_jobs = dp.create_jobs_from_excel_in(excel_jobs, active_sheet.max_column)
    return excel_jobs


def create_write_new_excel(new: list[dp.Job], old: list[dp.Job], old_path: string, new_path: string):
    max_col: int
    job_list: list
    title_row: list[str] = []

    if old_path != '':
        old_sheet = opxl.load_workbook(old_path).active

        job_list = dp.compare_jobs(new, old)
        max_col = old_sheet.max_column
        tr_gen = old_sheet.iter_rows(max_row=1, values_only=True)

        for cell in tr_gen:
            for val in cell:
                if val != 'Notes':
                    title_row.append(val)

    else:
        job_list = new
        max_col = 6
        title_row = ['Column1', 'Job No', 'Description', 'Column2', 'PM', 'Column5']

    formatted_job_list = dp.format_jobs_as_excel(job_list, max_col - 1)

    new_file = opxl.Workbook()
    sheet = new_file.active

    title_row.append(dp.pdf_date)
    title_row.append('Notes')
    sheet.append(title_row)

    for row_num, row in enumerate(formatted_job_list, 1):

        # keep track of if rows are grouped/hidden
        gp = False
        hd = False

        current_mod = (row_num - 1) % 4  # keeps track of first, second, third, fourth row of job

        # take any grouped/hidden statements out and act on them
        if 'Grouped=True' in row:
            gp = True
            row.remove('Grouped=True')
        if 'Hidden=True' in row:
            hd = True
            row.remove('Hidden=True')

        sheet.append(row)  # append cleaned up row

        if current_mod == 3 and gp is True:  # if last row of job, group/hide as needed
            sheet.row_dimensions.group(row_num - 2, row_num + 1, hidden=hd)

    # todo: append end stuff

    # formatting :3
    vf.format_widths(sheet)
    vf.format_sheet_style(sheet, '00FFFFFF', '00DDEBF7')

    new_file.save(new_path)

    return formatted_job_list
