
# excel sheet formatting stuff here

from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

def format_widths(sheet):
    # hardcoding column widths but can change this later
    sheet.column_dimensions['A'].width = 10.5
    sheet.column_dimensions['B'].width = 10.5
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 6
    sheet.column_dimensions['F'].width = 10

    max_col = sheet.max_column

    for col in range(7, max_col):
        sheet.column_dimensions[get_column_letter(col)].width = 12  # all the costs

    sheet.column_dimensions[get_column_letter(max_col)].width = 25  # notes column


def format_sheet_style(sheet, line_color1, line_color2):
    for i, row in enumerate(sheet.iter_rows()):
        for cell in row:
            # bold title row
            if i == 0:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # format cells with numbers so that negatives flip and turn red
            cell.number_format = '#,##0_);[Red](#,##0)'

            # color every other cell
            if i % 2 == 0:
                cell.fill = PatternFill(patternType='solid', fgColor=line_color1)
            else:
                cell.fill = PatternFill(patternType='solid', fgColor=line_color2)

            # border between all cells to see them easier
            cell.border = Border(
                left=Side(border_style='thin', color='002F75B5'),
                right=Side(border_style='thin', color='002F75B5'),
                top=Side(border_style='thin', color='002F75B5'),
                bottom=Side(border_style='thin', color='002F75B5')
            )

            # horizontal border between jobs
            if i % 4 == 0:
                cell.border = Border(
                    left=Side(border_style='thin', color='002F75B5'),
                    right=Side(border_style='thin', color='002F75B5'),
                    top=Side(border_style='thin', color='002F75B5'),
                    bottom=Side(border_style='medium', color='00000000')
                )
